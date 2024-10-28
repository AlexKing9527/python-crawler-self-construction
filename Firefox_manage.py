import openpyxl
import time
# import pic_ocr
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
import tkinter as tk

filepath = ''
# root = tk.Tk()
# root.title("LIMS非特结果导出文件的绝对地址，相关Excel所有内容均会上传，需自己把控所需上传内容")
# entry = tk.Entry(root, width=100)
# entry.pack()
# def on_click():
#     global filepath
#     filepath = entry.get()
#     entry.destroy()
#     root.destroy()
# button = tk.Button(root, text="确认", command=on_click)
# button.pack()
# root.mainloop()


# def search_col(name):
#     for cell in sheet[1]:
#         if cell.value == name:
#             return cell.column, get_column_letter(cell.column)

# print(filepath)
# wb = openpyxl.load_workbook(filepath)
# sheet = wb.active

# ftslbh = search_col('非特受理编号')
# fxx = search_col('分析项')
# jyjg = search_col('检验结果')
# jyff = search_col('检验方法')
# ffjcnd = search_col('方法检出浓度')
# dw = search_col('单位')
# xz = search_col('限值')

# # excel文件中需上传结果和附件的文件编号
# bhs = list({i.value for i in sheet[ftslbh[1]] if i.value not in ['非特受理编号', '']})

# 登录

# 无头
options = webdriver.FirefoxOptions()
options.add_argument("--headless")
options.add_argument("--disable-gpu")
driver = webdriver.Firefox(options=options)

# 有头
# driver = webdriver.Firefox()

driver.get(r'https://zwfw.nmpa.gov.cn/web/user/login')
driver.implicitly_wait(5)
driver.find_element(By.XPATH, '//*[@id="username-2"]').clear()
driver.find_element(By.XPATH, '//*[@id="username-2"]').send_keys('13631330114')
driver.find_element(By.XPATH, '//*[@id="password-2"]').send_keys('ZGCfeite123456')
while True:
    text = '0'
    while len(text) != 4:
        driver.find_element(By.XPATH, '//*[@id="rankings"]/div[3]/button').click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, '//*[@id="img_2"]').screenshot('login.png')
        # text = pic_ocr.character_recognition('login.png')
    driver.find_element(By.XPATH, '//*[@id="verifycode-2"]').clear()
    driver.find_element(By.XPATH, '//*[@id="verifycode-2"]').send_keys(text)
    driver.find_element(By.XPATH, '//*[@id="pane-first"]/div[2]/button').click()
    time.sleep(9)
    print(text)
    elements = driver.find_elements(By.XPATH, '//*[@id="pane-first"]/div[2]/button')
    if len(elements) == 0:
        break

# 进入录入界面
driver.find_element(By.XPATH, '//*[@id="main"]/div/div[1]/a[2]').click()
time.sleep(3)
driver.find_element(By.XPATH, '/html/body/div[6]/div/div[2]/div/ul/li[1]').click()
time.sleep(3)
driver.find_element(By.XPATH, '/html/body/div[13]/div[3]/a[1]').click()
time.sleep(3)
handl = driver.window_handles
driver.switch_to.window(handl[-1])
driver.switch_to.frame(driver.find_element(By.XPATH, '/html/body/div[3]/div[1]/iframe'))
time.sleep(1)
driver.find_element(By.XPATH, '//*[@id="opclose_10"]').click()
time.sleep(1)
driver.find_element(By.XPATH, '//*[@id="name_9cf61088306949c79585606b1963200e"]').click()
time.sleep(1)
driver.switch_to.default_content()

for bh in bhs:
    # 查询并填写结果
    driver.switch_to.frame(driver.find_element(By.XPATH, '//*[@id="mainframe"]'))
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="ename"]').clear()
    driver.find_element(By.XPATH, '//*[@id="ename"]').send_keys(bh)
    driver.find_element(By.XPATH, '//*[@id="search"]').click()
    time.sleep(1)
    driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/table/tbody[2]/tr/td[8]/a').click()
    time.sleep(3)
    driver.switch_to.window(handl[-1])

    js = 'document.querySelector("#commitJg1").click();'
    driver.execute_script(js)
    driver.switch_to.frame(driver.find_element(By.XPATH, '//*[@id="winbox_frame"]'))

    # 微生物结果填写
    time.sleep(3)
    a = driver.find_elements(By.XPATH, '/html/body/dl[1]/dd/div[1]')
    b = len(a)
    c = []
    for i in range(1, b):
        for name in sheet[fxx[1]]:
            if driver.find_element(By.CSS_SELECTOR,
                                   f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_2').text == name.value and \
                    sheet[f'{ftslbh[1]}{name.row}'].value == bh:
                c.append(name.value)
                # 结果
                driver.find_element(By.CSS_SELECTOR,
                                    f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_3 div#e_e{i}.appsr.divInput').clear()
                driver.find_element(By.CSS_SELECTOR,
                                    f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_3 div#e_e{i}.appsr.divInput').send_keys(
                    sheet[f'{jyjg[1]}{name.row}'].value)
                # 单位
                driver.find_element(By.CSS_SELECTOR,
                                    f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_5 input#e_unit_e{i}.appsr').clear()
                driver.find_element(By.CSS_SELECTOR,
                                    f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_5 input#e_unit_e{i}.appsr').send_keys(
                    sheet[f'{dw[1]}{name.row}'].value)
                # 限值
                # driver.find_element(By.CSS_SELECTOR, f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_6 div#e_e{i}.appsr.divInput').clear()
                # driver.find_element(By.CSS_SELECTOR, f'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e{i}.jyxm_e div.appcomp_6 div#e_e{i}.appsr.divInput').send_keys(sheet[f'{xz[1]}{name.row}'].value)
                break
    # 微生物结论填写
    driver.find_element(By.CSS_SELECTOR,
                        'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e6.jyxm_e div.appcomp_6 div.divInput.appcompsr').clear()
    driver.find_element(By.CSS_SELECTOR,
                        'html body#body dl#jyxm_e.appcomp.jyxm_appcomp dd#jyxm_e6.jyxm_e div.appcomp_6 div.divInput.appcompsr').send_keys(
        '、'.join(c) + '检验通过')

    # 保存结果
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="btn_jg_bc"]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '/html/body/div[4]/div[1]/div[2]/div[2]/div[1]/div/div/div/div[2]/div[4]/input[1]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '/html/body/div[4]/div[1]/div[2]/div[2]/div[1]/div/div/div/div[2]/div[4]/input').click()
    time.sleep(1)

    # 理化结果填写
    driver.find_element(By.XPATH, '//*[@id="tlTab_f"]').click()
    table = driver.find_element(By.ID, 'jyxm_f100')
    rows = table.find_elements(By.TAG_NAME, 'tr')
    d = []
    for i in range(1, len(rows)):
        colums = rows[i].find_elements(By.TAG_NAME, 'td')
        for name in sheet[fxx[1]]:
            if colums[0].text == name.value and sheet[f'{ftslbh[1]}{name.row}'].value == bh and name.value not in d:
                d.append(name.value)
                # 单位
                colums[1].find_element(By.TAG_NAME, 'input').clear()
                colums[1].find_element(By.TAG_NAME, 'input').send_keys(sheet[f'{dw[1]}{name.row}'].value)
                # 结果
                colums[2].find_element(By.TAG_NAME, 'div').clear()
                colums[2].find_element(By.TAG_NAME, 'div').send_keys(sheet[f'{jyjg[1]}{name.row}'].value)
                # 检验方法
                colums[3].find_element(By.TAG_NAME, 'input').clear()
                colums[3].find_element(By.TAG_NAME, 'input').send_keys(sheet[f'{jyff[1]}{name.row}'].value)
                # 方法检出浓度
                colums[4].find_element(By.TAG_NAME, 'input').clear()
                colums[4].find_element(By.TAG_NAME, 'input').send_keys(sheet[f'{ffjcnd[1]}{name.row}'].value)
                break
            else:
                colums[1].find_element(By.TAG_NAME, 'input').clear()
                colums[2].find_element(By.TAG_NAME, 'div').clear()
                colums[3].find_element(By.TAG_NAME, 'input').clear()
                colums[4].find_element(By.TAG_NAME, 'input').clear()
    # 结论
    rows[-1].find_elements(By.TAG_NAME, 'td')[1].find_element(By.TAG_NAME, 'div').clear()
    rows[-1].find_elements(By.TAG_NAME, 'td')[1].find_element(By.TAG_NAME, 'div').send_keys('、'.join(d) + '检验通过')
    # 保存结果
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="btn_jg_bc"]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '/html/body/div[4]/div[1]/div[2]/div[2]/div[1]/div/div/div/div[2]/div[4]/input[1]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '/html/body/div[4]/div[1]/div[2]/div[2]/div[1]/div/div/div/div[2]/div[4]/input').click()
    time.sleep(1)

    # 退出界面
    driver.switch_to.default_content()
    js = 'document.querySelector("#winbox_cont > div:nth-child(1) > div:nth-child(2)").click();'
    driver.execute_script(js)

    """ 
    #上传附件 (未安装flash，无法上传）
    file_input = driver.find_element(By.ID, "attach1")
    file_input.click()
    driver.find_element(By.XPATH,'//*[@id="jq_alert_window_box"]/div/div/div[2]/div[4]/input[1]').click()
    time.sleep(2)
    driver.switch_to.frame(driver.find_element(By.XPATH,'//*[@id="cboxLoadedContent"]/iframe'))
    dd = driver.find_element(By.XPATH,'//*[@id="file_upload"]')
    dd.send_keys("C:\\Users\\53046\Desktop\新建文件夹\XRD衍射仪的工作原理(1).ppt")
    driver.find_element(By.XPATH,'/html/body/div[1]/div[2]').click()
    input('queren')
    """

    # 返回主页
    js = 'document.querySelector("#btnToBack").click();'
    driver.execute_script(js)
    driver.switch_to.window(handl[-1])

# driver.switch_to.alert.accept() #接受弹窗
driver.quit()
